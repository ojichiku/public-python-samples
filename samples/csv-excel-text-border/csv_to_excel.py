from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter


USAGE = "使い方: python csv_to_excel.py 入力ファイル.csv"
MAX_COLUMN_WIDTH = 40
MIN_COLUMN_WIDTH = 8


def read_csv_rows(csv_path: Path) -> list[list[str]]:
    """CSV file contents as rows of strings."""
    with csv_path.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.reader(csv_file)
        return [[str(value) for value in row] for row in reader]


def write_excel(rows: list[list[str]], output_path: Path) -> None:
    """Write CSV rows to an Excel file without changing values."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CSV"

    max_columns = max(len(row) for row in rows)
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    for row_index, row in enumerate(rows, start=1):
        for column_index in range(1, max_columns + 1):
            value = row[column_index - 1] if column_index <= len(row) else ""
            cell = worksheet.cell(
                row=row_index,
                column=column_index,
                value=str(value),
            )
            cell.number_format = "@"
            cell.border = thin_border
            if row_index == 1:
                cell.font = Font(bold=True)

    adjust_column_widths(worksheet, rows, max_columns)
    workbook.save(output_path)


def adjust_column_widths(worksheet, rows: list[list[str]], max_columns: int) -> None:
    """Set readable column widths based on CSV contents."""
    for column_index in range(1, max_columns + 1):
        max_length = 0
        for row in rows:
            value = row[column_index - 1] if column_index <= len(row) else ""
            max_length = max(max_length, len(str(value)))

        width = min(max(max_length + 2, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def convert_csv_to_excel(csv_path: Path) -> Path:
    """Convert a CSV file to an Excel file in the same directory."""
    rows = read_csv_rows(csv_path)
    if not rows:
        raise ValueError("CSVファイルにデータがありません。")

    output_path = csv_path.with_suffix(".xlsx")
    write_excel(rows, output_path)
    return output_path


def validate_input(args: list[str]) -> Path | None:
    """Validate command line arguments and return the CSV path."""
    if not args:
        print(USAGE)
        return None

    csv_path = Path(args[0])

    if csv_path.suffix.lower() != ".csv":
        print("エラー: .csv ファイルを指定してください。", file=sys.stderr)
        return None

    if not csv_path.exists():
        print(
            f"エラー: 指定されたCSVファイルが見つかりません: {csv_path}",
            file=sys.stderr,
        )
        return None

    if not csv_path.is_file():
        print(f"エラー: ファイルを指定してください: {csv_path}", file=sys.stderr)
        return None

    return csv_path


def main() -> int:
    csv_path = validate_input(sys.argv[1:])
    if csv_path is None:
        return 1

    try:
        output_path = convert_csv_to_excel(csv_path)
    except Exception as exc:
        print(f"エラー: 変換に失敗しました。{exc}", file=sys.stderr)
        return 1

    print(f"Excelファイルを作成しました: {output_path.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
